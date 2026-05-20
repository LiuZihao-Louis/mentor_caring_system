from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def read_xlsx_as_dicts(path: str) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() for h in rows[0]]
    result: list[dict] = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        result.append({headers[i]: row[i] for i in range(len(headers))})
    return result
